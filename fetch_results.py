"""
fetch_results.py
----------------
Fetches completed World Cup 2026 match results from ESPN's unofficial API
and writes home/away scores into columns D and E of the FixturesResults sheet
in World Cup 2026 Sweeps.xlsx.

Run manually:  python3 fetch_results.py
Scheduled:     GitHub Actions (.github/workflows/daily_update.yml)
"""

import json
import os
import urllib.request
from datetime import date, datetime, timedelta
import openpyxl

# ── Config ────────────────────────────────────────────────────────────────────
XLSX_PATH = os.environ.get(
    "XLSX_PATH",
    os.path.join(os.path.dirname(__file__), "World Cup 2026 Sweeps.xlsx"),
)
TOURNAMENT_START = date(2026, 6, 11)
TOURNAMENT_END   = date(2026, 7, 19)

# ── Team name mapping (ESPN displayName → sheet name) ─────────────────────────
ESPN_TO_SHEET = {
    "Bosnia-Herzegovina": "Bosnia and Herzegovina",
    "Congo DR":           "DR Congo",
    "Türkiye":            "Turkey",
    "United States":      "USA",
}

# ── ESPN API helpers ──────────────────────────────────────────────────────────
BASE_URL = "https://site.api.espn.com/apis/site/v2/sports/soccer/fifa.world/scoreboard"

def fetch_matches_for_date(d: date) -> list[dict]:
    """
    Returns all matches for the given date (completed and upcoming).
    Completed matches include scores; upcoming matches have scores as None.
    """
    url = f"{BASE_URL}?dates={d.strftime('%Y%m%d')}"
    try:
        with urllib.request.urlopen(url, timeout=15) as resp:
            data = json.load(resp)
    except Exception as e:
        print(f"  Warning: could not fetch {d} — {e}")
        return []

    results = []
    for event in data.get("events", []):
        comp = event["competitions"][0]
        status = comp["status"]["type"]["name"]  # e.g. "STATUS_FINAL", "STATUS_SCHEDULED"
        finished = status in ("STATUS_FINAL", "STATUS_FULL_TIME",
                              "STATUS_FINAL_PEN", "STATUS_FINAL_AET")

        home_comp = next((t for t in comp["competitors"] if t["homeAway"] == "home"), None)
        away_comp = next((t for t in comp["competitors"] if t["homeAway"] == "away"), None)
        if not home_comp or not away_comp:
            continue

        home_name = ESPN_TO_SHEET.get(home_comp["team"]["displayName"], home_comp["team"]["displayName"])
        away_name = ESPN_TO_SHEET.get(away_comp["team"]["displayName"], away_comp["team"]["displayName"])

        home_score = None
        away_score = None
        if finished:
            try:
                home_score = int(home_comp["score"])
                away_score = int(away_comp["score"])
            except (KeyError, ValueError, TypeError):
                continue

        # Penalty shootout scores (only present for STATUS_FINAL_PEN)
        home_pen = home_comp.get("shootoutScore")
        away_pen = away_comp.get("shootoutScore")
        if home_pen is not None:
            home_pen = int(home_pen)
        if away_pen is not None:
            away_pen = int(away_pen)

        results.append({
            "home":       home_name,
            "away":       away_name,
            "home_score": home_score,
            "away_score": away_score,
            "finished":   finished,
            "home_winner": home_comp.get("winner", None),
            "away_winner": away_comp.get("winner", None),
            "home_pen":   home_pen,
            "away_pen":   away_pen,
            "event_id":   event.get("id"),
            "home_team_id": home_comp["team"].get("id"),
        })

    return results


# ── Main update logic ─────────────────────────────────────────────────────────
def main():
    print(f"Loading: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["FixturesResults"]

    # Build index: (home_team, away_team) → row number
    fixture_index: dict[tuple[str, str], int] = {}
    # Secondary index: date → list of rows with no team names (knockout placeholders)
    empty_rows_by_date: dict[date, list[int]] = {}
    for row in range(3, 200):
        raw_date = ws.cell(row, 1).value
        if not raw_date:
            break
        home = ws.cell(row, 3).value  # col C
        away = ws.cell(row, 6).value  # col F
        # Treat placeholder names (e.g. "Round of 32 3 Winner") as empty/fillable
        is_placeholder = lambda v: not v or (isinstance(v, str) and "Winner" in v)
        if home and away and not is_placeholder(home) and not is_placeholder(away):
            fixture_index[(home, away)] = row
        elif raw_date and (is_placeholder(home) or is_placeholder(away)):
            d = raw_date.date() if isinstance(raw_date, datetime) else raw_date
            empty_rows_by_date.setdefault(d, []).append(row)

    print(f"Indexed {len(fixture_index)} fixtures in the sheet.")
    empty_count = sum(len(v) for v in empty_rows_by_date.values())
    if empty_count:
        print(f"Found {empty_count} empty knockout placeholder rows.")

    # Fetch all matches from tournament start through end (scores + upcoming fixtures)
    fetch_date = TOURNAMENT_START
    updated = 0
    skipped = 0
    filled = 0

    while fetch_date <= TOURNAMENT_END:
        matches = fetch_matches_for_date(fetch_date)
        completed = [m for m in matches if m["finished"]]
        upcoming  = [m for m in matches if not m["finished"]]
        if completed:
            print(f"  {fetch_date}: {len(completed)} completed match(es)")
        if upcoming:
            print(f"  {fetch_date}: {len(upcoming)} upcoming match(es)")

        for m in matches:
            key = (m["home"], m["away"])

            if key in fixture_index:
                row = fixture_index[key]
            elif fetch_date in empty_rows_by_date and empty_rows_by_date[fetch_date]:
                # Knockout placeholder: assign team names to the first available empty row
                row = empty_rows_by_date[fetch_date].pop(0)
                ws.cell(row, 1).value = datetime(fetch_date.year, fetch_date.month, fetch_date.day)
                ws.cell(row, 3).value = m["home"]
                ws.cell(row, 6).value = m["away"]
                fixture_index[key] = row
                print(f"    Filled knockout row {row}: {m['home']} vs {m['away']}")
                filled += 1
            else:
                print(f"    !! No row found for: {m['home']} vs {m['away']}")
                skipped += 1
                continue

            # Only write scores for completed matches
            if not m["finished"]:
                continue
            # Always write event ID if available
            if m.get("event_id") and not ws.cell(row, 12).value:
                ws.cell(row, 12).value = m["event_id"]
                ws.cell(row, 13).value = m["home_team_id"]
                filled += 1  # ensure save happens

            existing_d = ws.cell(row, 4).value
            existing_e = ws.cell(row, 5).value
            existing_hp = ws.cell(row, 8).value
            existing_ap = ws.cell(row, 9).value
            scores_match = (existing_d == m["home_score"] and existing_e == m["away_score"])
            pens_match = (existing_hp == m["home_pen"] and existing_ap == m["away_pen"])
            if scores_match and pens_match:
                continue  # already up to date
            ws.cell(row, 4).value = m["home_score"]
            ws.cell(row, 5).value = m["away_score"]
            if m["home_pen"] is not None:
                ws.cell(row, 8).value = m["home_pen"]
                ws.cell(row, 9).value = m["away_pen"]
            print(f"    Updated row {row}: {m['home']} {m['home_score']}-{m['away_score']} {m['away']}" +
                  (f" (pens {m['home_pen']}-{m['away_pen']})" if m["home_pen"] is not None else ""))
            updated += 1

        fetch_date += timedelta(days=1)

    if updated > 0 or filled > 0:
        wb.save(XLSX_PATH)
        print(f"\nSaved. {updated} score(s) written, {filled} knockout fixture(s) filled.")
    else:
        print(f"\nNo changes needed. ({skipped} unmatched)")


if __name__ == "__main__":
    main()
