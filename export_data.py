"""
export_data.py
--------------
Reads World Cup 2026 Sweeps.xlsx and exports data.json for the website.
Recalculates all points from raw scores — does not depend on cached formula values.

Run:  python3 export_data.py
      XLSX_PATH=/path/to/file.xlsx python3 export_data.py
"""

import json
import os
import urllib.request
from datetime import datetime, date, timezone, timedelta
import openpyxl

BST = timezone(timedelta(hours=1))

def fetch_kickoff_times() -> dict[tuple[str, str], str]:
    """
    Fetches kick-off times from ESPN for all tournament dates.
    Returns a dict of (home_team, away_team) -> "HH:MM BST" string.
    Uses ESPN_TO_SHEET mapping to normalise team names.
    """
    ESPN_TO_SHEET = {
        "Bosnia-Herzegovina": "Bosnia and Herzegovina",
        "Congo DR":           "DR Congo",
        "Türkiye":            "Turkey",
        "United States":      "USA",
    }
    BASE = "https://site.api.espn.com/apis/site/v2/sports/soccer/fifa.world/scoreboard"
    times = {}
    fetch_date = date(2026, 6, 11)
    end_date   = date(2026, 7, 20)
    while fetch_date <= end_date:
        url = f"{BASE}?dates={fetch_date.strftime('%Y%m%d')}"
        try:
            with urllib.request.urlopen(url, timeout=10) as r:
                data = json.load(r)
        except Exception:
            fetch_date += timedelta(days=1)
            continue
        for event in data.get("events", []):
            raw = event.get("date")  # e.g. "2026-06-11T19:00Z"
            if not raw:
                continue
            dt_utc = datetime.strptime(raw, "%Y-%m-%dT%H:%MZ").replace(tzinfo=timezone.utc)
            dt_bst = dt_utc.astimezone(BST)
            time_str = dt_bst.strftime("%H:%M")
            comp = event["competitions"][0]
            home_c = next((t for t in comp["competitors"] if t["homeAway"] == "home"), None)
            away_c = next((t for t in comp["competitors"] if t["homeAway"] == "away"), None)
            if not home_c or not away_c:
                continue
            home = ESPN_TO_SHEET.get(home_c["team"]["displayName"], home_c["team"]["displayName"])
            away = ESPN_TO_SHEET.get(away_c["team"]["displayName"], away_c["team"]["displayName"])
            times[(home, away)] = time_str
        fetch_date += timedelta(days=1)
    return times

def fetch_shootout_plays(event_id: str, home_team_id: str) -> dict:
    """
    Fetches kick-by-kick penalty shootout data from ESPN play-by-play API.
    Returns {"home_kicks": [True/False, ...], "away_kicks": [True/False, ...]}
    where True = scored, False = missed/saved. Kicks are in order taken.
    """
    PLAYS_BASE = f"https://sports.core.api.espn.com/v2/sports/soccer/leagues/fifa.world/events/{event_id}/competitions/{event_id}/plays"
    try:
        # First get page count with a reasonable limit
        with urllib.request.urlopen(f"{PLAYS_BASE}?limit=200", timeout=10) as r:
            meta = json.load(r)
        page_count = meta.get("pageCount", 1)
        # Shootout plays are on the last page
        with urllib.request.urlopen(f"{PLAYS_BASE}?limit=200&page={page_count}", timeout=10) as r:
            data = json.load(r)
    except Exception as e:
        print(f"  Warning: could not fetch shootout plays for event {event_id} — {e}")
        return None

    home_kicks = []
    away_kicks = []
    for item in data.get("items", []):
        if not item.get("shootout"):
            continue
        scored = item.get("scoringPlay", False)
        # Determine which team took the kick from participants
        team_ref = ""
        for p in item.get("participants", []):
            if p.get("type") in ("shooter", "scorer"):
                team_ref = p.get("team", {}).get("$ref", "")
                break
        is_home = f"teams/{home_team_id}" in team_ref
        if is_home:
            home_kicks.append(scored)
        else:
            away_kicks.append(scored)

    if not home_kicks and not away_kicks:
        return None
    return {"home_kicks": home_kicks, "away_kicks": away_kicks}


XLSX_PATH = os.environ.get(
    "XLSX_PATH",
    os.path.join(os.path.dirname(__file__), "World Cup 2026 Sweeps.xlsx"),
)
OUT_PATH = os.path.join(os.path.dirname(__file__), "data.json")

# Stage labels by date ranges
STAGE_RANGES = [
    (date(2026, 6, 11), date(2026, 6, 27), "Group Stage"),
    (date(2026, 6, 28), date(2026, 7,  3), "Round of 32"),
    (date(2026, 7,  4), date(2026, 7,  7), "Round of 16"),
    (date(2026, 7,  9), date(2026, 7, 11), "Quarter-finals"),
    (date(2026, 7, 14), date(2026, 7, 15), "Semi-finals"),
    (date(2026, 7, 18), date(2026, 7, 18), "Third Place"),
    (date(2026, 7, 19), date(2026, 7, 19), "Final"),
]

def get_stage(d: date) -> str:
    for start, end, label in STAGE_RANGES:
        if start <= d <= end:
            return label
    return "TBD"


def calc_points(home_score: int, away_score: int):
    """Returns (home_points, away_points) given a scoreline."""
    if home_score == away_score:
        return 1, 1
    elif home_score > away_score:
        return 3 + home_score - away_score, away_score - home_score
    else:
        return home_score - away_score, 3 + away_score - home_score


def main():
    print(f"Reading: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    print("Fetching kick-off times from ESPN...")
    kickoff_times = fetch_kickoff_times()
    print(f"  Got {len(kickoff_times)} kick-off times.")

    # ── 1. Player → teams mapping ────────────────────────────────────────────
    ws_at = wb["All Teams"]
    players_order = [ws_at.cell(1, c).value for c in range(1, 17) if ws_at.cell(1, c).value]
    team_to_player = {}
    player_teams = {}
    for col_idx, player in enumerate(players_order, start=1):
        teams = []
        for row in range(2, 5):
            t = ws_at.cell(row, col_idx).value
            if t:
                teams.append(t)
                team_to_player[t] = player
        player_teams[player] = teams

    # ── 2. Fixtures + scores ─────────────────────────────────────────────────
    ws_fr = wb["FixturesResults"]
    fixtures = []
    for row in range(3, 200):
        raw_date = ws_fr.cell(row, 1).value
        home     = ws_fr.cell(row, 3).value
        away     = ws_fr.cell(row, 6).value
        if not raw_date:
            break
        # Include KO placeholders (TBD teams) so the website shows all stages

        d = raw_date.date() if isinstance(raw_date, datetime) else raw_date
        home_score = ws_fr.cell(row, 4).value  # col D
        away_score = ws_fr.cell(row, 5).value  # col E
        home_pen   = ws_fr.cell(row, 8).value  # col H: home penalty score
        away_pen   = ws_fr.cell(row, 9).value  # col I: away penalty score
        event_id     = ws_fr.cell(row, 12).value  # col L: ESPN event ID
        home_team_id = ws_fr.cell(row, 13).value  # col M: ESPN home team ID

        played = (home_score is not None and away_score is not None)
        home_name = home or "TBD"
        away_name = away or "TBD"
        kickoff = kickoff_times.get((home_name, away_name))

        # Fetch kick-by-kick shootout data for penalty matches
        shootout = None
        if home_pen is not None and event_id and home_team_id:
            print(f"  Fetching shootout plays for {home_name} vs {away_name}...")
            shootout = fetch_shootout_plays(str(event_id), str(home_team_id))

        fixtures.append({
            "date":        d.isoformat(),
            "kickoff_bst": kickoff,  # e.g. "20:00", or null if unknown
            "home":        home_name,
            "away":        away_name,
            "home_score":  int(home_score) if played else None,
            "away_score":  int(away_score) if played else None,
            "home_pen":    int(home_pen) if home_pen is not None else None,
            "away_pen":    int(away_pen) if away_pen is not None else None,
            "shootout":    shootout,  # {"home_kicks": [T/F,...], "away_kicks": [T/F,...]}
            "status":      "played" if played else ("live" if d == date.today() else "upcoming"),
            "stage":       get_stage(d),
        })

    # ── 3. Per-team stats ─────────────────────────────────────────────────────
    team_stats = {t: {"played": 0, "won": 0, "lost": 0, "drawn": 0,
                      "goals_for": 0, "goals_against": 0, "points": 0, "eliminated": False}
                  for t in team_to_player}

    for fx in fixtures:
        if fx["home_score"] is None:
            continue
        home, away = fx["home"], fx["away"]
        # Include penalty goals in the total for points calculation
        hs = fx["home_score"] + (fx["home_pen"] or 0)
        as_ = fx["away_score"] + (fx["away_pen"] or 0)
        hp, ap = calc_points(hs, as_)

        # Goals for/against include penalty goals
        for team, gf, ga, pts in [(home, hs, as_, hp), (away, as_, hs, ap)]:
            if team not in team_stats:
                continue
            s = team_stats[team]
            s["played"]       += 1
            s["goals_for"]    += gf
            s["goals_against"] += ga
            s["points"]       += pts
            if gf > ga:
                s["won"]  += 1
            elif gf < ga:
                s["lost"] += 1
            else:
                s["drawn"] += 1

    # Mark teams eliminated in knockout stages (loser of a played KO match)
    for fx in fixtures:
        if fx["home_score"] is None or fx["stage"] == "Group Stage":
            continue
        hs = fx["home_score"] + (fx["home_pen"] or 0)
        as_ = fx["away_score"] + (fx["away_pen"] or 0)
        if hs > as_:
            loser = fx["away"]
        elif as_ > hs:
            loser = fx["home"]
        else:
            loser = None
        if loser and loser in team_stats:
            team_stats[loser]["eliminated"] = True

    # Mark teams eliminated in the group stage (not in any knockout fixture)
    ko_teams = set()
    has_ko_fixtures = False
    for fx in fixtures:
        if fx["stage"] != "Group Stage":
            if fx["home"] != "TBD":
                ko_teams.add(fx["home"])
            if fx["away"] != "TBD":
                ko_teams.add(fx["away"])
            has_ko_fixtures = True
    if has_ko_fixtures:
        for team in team_stats:
            if team not in ko_teams and not team_stats[team]["eliminated"]:
                team_stats[team]["eliminated"] = True

    # ── 4. Per-player aggregation ─────────────────────────────────────────────
    player_data = []
    for player in players_order:
        teams = player_teams[player]
        breakdown = []
        total_points = 0
        for t in teams:
            s = team_stats.get(t, {})
            pts = s.get("points", 0)
            total_points += pts
            gf = s.get("goals_for", 0)
            ga = s.get("goals_against", 0)
            breakdown.append({
                "team":          t,
                "played":        s.get("played", 0),
                "won":           s.get("won", 0),
                "lost":          s.get("lost", 0),
                "drawn":         s.get("drawn", 0),
                "goals_for":     gf,
                "goals_against": ga,
                "goal_diff":     gf - ga,
                "points":        pts,
                "eliminated":    s.get("eliminated", False),
            })
        player_data.append({
            "name":         player,
            "teams":        teams,
            "total_points": total_points,
            "breakdown":    breakdown,
        })

    # Sort by points desc, then name
    player_data.sort(key=lambda p: (-p["total_points"], p["name"]))
    for rank, p in enumerate(player_data, 1):
        p["rank"] = rank

    # ── 5. Daily cumulative scores ────────────────────────────────────────────
    # For each match date (sorted), compute how many points each player earned
    # that day, then build running cumulative totals.
    from collections import defaultdict

    # Group played fixtures by date
    points_by_date: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for fx in fixtures:
        if fx["home_score"] is None:
            continue
        d = fx["date"]
        hs = fx["home_score"] + (fx["home_pen"] or 0)
        as_ = fx["away_score"] + (fx["away_pen"] or 0)
        hp, ap = calc_points(hs, as_)
        for team, pts in [(fx["home"], hp), (fx["away"], ap)]:
            owner = team_to_player.get(team)
            if owner:
                points_by_date[d][owner] += pts

    sorted_dates = sorted(points_by_date.keys())
    all_players = [p["name"] for p in player_data]

    # Build cumulative series
    cumulative = {name: 0 for name in all_players}
    daily_series: dict[str, list[int]] = {name: [] for name in all_players}
    for d in sorted_dates:
        day_pts = points_by_date[d]
        for name in all_players:
            cumulative[name] += day_pts.get(name, 0)
            daily_series[name].append(cumulative[name])

    daily_scores = {
        "dates":  sorted_dates,
        "series": [{"name": name, "data": daily_series[name]} for name in all_players],
    }

    # ── 6. Write output ───────────────────────────────────────────────────────
    output = {
        "last_updated": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "players":      player_data,
        "fixtures":     fixtures,
        "daily_scores": daily_scores,
    }

    with open(OUT_PATH, "w") as f:
        json.dump(output, f, indent=2)

    played_count = sum(1 for fx in fixtures if fx["status"] == "played")
    print(f"Exported {len(player_data)} players, {len(fixtures)} fixtures ({played_count} played) → {OUT_PATH}")


if __name__ == "__main__":
    main()
